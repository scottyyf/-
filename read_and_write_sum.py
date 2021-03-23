#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
File: read_and_write_sum.py.py
Author: Scott Yang(Scott)
Email: yangyingfa@skybility.com
Copyright: Copyright (c) 2021, Skybility Software Co.,Ltd. All rights reserved.
Description:
"""
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class Read(object):
    def __init__(self, xls_file_path='xxx.xlsx'):
        self.xls_file_path = xls_file_path
        self.wb = None

    def load_and_get_wb(self):
        if not os.path.isfile(self.xls_file_path):
            raise TypeError(f'{self.xls_file_path} is not a exist file')

        try:
            self.wb = load_workbook(self.xls_file_path)

        except Exception as e:
            raise e

    def parse_to_data(self):
        # total_data = {'table_name': [[row1], [row2], ....]}
        total_data = {}
        self.load_and_get_wb()
        all_tables = self.wb.sheetnames
        for table in all_tables:
            table_data = []
            sheet = self.wb[table]
            max_row = sheet.max_row
            max_column = sheet.max_column

            for row in range(1, max_row + 1):
                row_data = []

                for column in range(1, max_column + 1):
                    _column = get_column_letter(column)
                    value = sheet[f'{_column}{row}'].value
                    if max_column > 1:
                        _v = sheet[f'B{row}'].value
                        if _v and isinstance(_v, int) and _v <= 1000:
                            break

                    if value == '位置描叙' or not value:
                        break

                    row_data.append(value)

                if row_data:
                    table_data.append(row_data)

            if table_data:
                table_data = sorted(table_data, key=lambda x: x[1])

            if len(table_data) > 4:
                table_data = table_data[2:-2]

            total_data[table] = table_data

        return total_data

    def statistic_table(self):
        # work_sheet = wb.create_sheet('统计表', 0)
        total_data = self.parse_to_data()
        ret = []
        for k, v in total_data.items():
            highest = 0
            lowest = 0
            pingjun = 0
            per_mi_price = 0
            if v:
                highest = v[-1][1]
                lowest = v[0][1]
                sum = 0
                sum_mi = 0
                for s in v:
                    sum += s[1]
                    sum_mi += s[2]

                pingjun = sum//len(v)
                per_mi_price = sum // sum_mi

            statistic_data = StatisticData(k, highest, lowest, pingjun, per_mi_price)
            ret.append(statistic_data)

        return ret

    def write_statistic_data(self, table_name='统计表'):
        ret = self.statistic_table()
        work_sheet = self.wb.create_sheet(table_name, 0)
        work_sheet['A1'] = '统计表来源'
        work_sheet['B1'] = '该区域最高价'
        work_sheet['C1'] = '该区域最低价'
        work_sheet['D1'] = '该区域平均价格'
        work_sheet['F1'] = '该区域每平价格'

        start = 2
        for static_zone_data in ret:
            work_sheet[f'A{start}'] = static_zone_data.table_name
            work_sheet[f'B{start}'] = static_zone_data.hightest_price
            work_sheet[f'C{start}'] = static_zone_data.lowest_price
            work_sheet[f'D{start}'] = static_zone_data.pingjun_price
            work_sheet[f'F{start}'] = static_zone_data.per_mi_price
            start += 1

        self.wb.save('安居客.xlsx')


class StatisticData(object):
    def __init__(self, table_name, hightest_rice, lowest_price, pingjun_price, per_mi_price):
        self.table_name = table_name
        self.hightest_price = hightest_rice
        self.lowest_price = lowest_price

        self.pingjun_price = pingjun_price
        self.per_mi_price = per_mi_price


if __name__ == '__main__':
    rd = Read()
    data = rd.write_statistic_data()
